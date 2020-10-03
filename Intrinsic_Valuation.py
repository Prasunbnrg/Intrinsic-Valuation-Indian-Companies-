''' Code to get the Intrinsic Value of a Company
    Takes file path with input data and assumptions
    Returns Value of the firm and financial projections '''

from urllib.request import urlopen
import string
import pandas as pd
import numpy as np
from bs4 import BeautifulSoup
import yahoo_fin.stock_info as si
from openpyxl import load_workbook

def importing_input_data(file_path):
    ''' Import data from input sheet of Intrinsic Valuation
        Return input data in DataFrames '''

    col1 = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]
    df1 = pd.read_excel(file_path, sheet_name='Input Sheet',
                        header=2, usecols=col1)
    df_f = df1.loc[0]

    col2 = [17, 18, 19, 20]
    df2 = pd.read_excel(file_path, sheet_name='Input Sheet',
                        header=2, index_col='Country', usecols=col2)
    df_rev = pd.DataFrame(df2).dropna()

    col3 = [21, 22, 23, 24, 25, 26, 27]
    df3 = pd.read_excel(file_path, sheet_name='Input Sheet',
                        header=2, index_col='Period', usecols=col3)
    df_a = pd.DataFrame(df3)

    input_data = [df_f, df_a, df_rev]
    return input_data

def get_financial_statements(url_company):
    ''' Import financial data from Money Control
        Return financial data in DataFrames '''

    fin_statements = []
    for i in url_company:
        html = urlopen(i)
        soup = BeautifulSoup(html, 'lxml')

        soup.find_all('tr')
        rows = soup.find_all('tr')

        for row in rows:
            row_td = row.find_all('td')

        str_cells = str(row_td)

        import re
        list_rows = []
        for row in rows:
            cells = row.find_all('td')
            str_cells = str(cells)
            clean = re.compile('<.*?>')
            clean2 = (re.sub(clean, '', str_cells))
            list_rows.append(clean2)

        df_scrape_data = pd.DataFrame(list_rows)
        df1 = df_scrape_data[0].str.split(', ', expand=True)
        df1[1].replace(',', '')
        df1[0] = df1[0].str.strip('[')
        df1.loc[1, 0] = 'Items'
        df1.columns = df1.loc[1]
        df1 = df1.drop([0, 1, 2])
        df1.index = df1['Items']
        df1 = df1.drop(['Items'], axis=1)
        df1 = df1.iloc[:, :-2]

        fin_statements.append(df1)
    return fin_statements

def replacecomma(var):
    ''' Replace comma (,) in financial numbers '''

    import re
    return float(re.sub(',', '', var))

def get_equity_research_data(df_bs, df_pl, df_cf, capex, statement, filename):
    ''' Takes financial statements
        Return necessary financial data for Intrinsic Valuation '''

    df1 = df_bs[[True if m in ['Inventories', 'Trade Receivables', 'Trade Payables',
                               'Short Term Borrowings', 'Long Term Borrowings',
                               'Total Capital And Liabilities', 'Total Shareholders Funds']
                 else False for m in df_bs.index.tolist()]]
    df1 = df1.applymap(replacecomma)

    df2 = df_pl[[True if m in ['Total Revenue', 'Cost Of Materials Consumed',
                               'Operating And Direct Expenses', 'Employee Benefit Expenses',
                               'Depreciation And Amortisation Expenses', 'Finance Costs',
                               'Other Expenses', 'Total Expenses', 'Total Tax Expenses',
                               'Profit/Loss Before Tax', 'Profit/Loss For The Period',
                               'Minority Interest']
                 else False for m in df_pl.index.tolist()]]
    df2 = df2.applymap(replacecomma)

    df3 = df_cf[[True if m in ['Net CashFlow From Operating Activities',
                               'Cash And Cash Equivalents End Of Year']
                 else False for m in df_cf.index.tolist()]]
    df3 = df3.applymap(replacecomma)

    df_column = df1.columns.tolist()
    capex.columns = df_column

    df = pd.concat([df1, df2, df3, capex], axis=0, sort=False)

    if statement == 1:
        df_val_data = \
            pd.DataFrame([df.loc['Inventories'], df.loc['Trade Receivables'],
                          df.loc['Trade Payables'],
                          (df.loc['Inventories'] + df.loc['Trade Receivables']
                           - df.loc['Trade Payables']) -
                          (df.loc['Inventories'].shift(-1) + df.loc['Trade Receivables'].shift(-1)
                           - df.loc['Trade Payables'].shift(-1)),
                          df.loc['Long Term Borrowings'], df.loc['Short Term Borrowings'],
                          df.loc['Long Term Borrowings'] + df.loc['Short Term Borrowings'],
                          (df.loc['Total Capital And Liabilities'] -
                           df.loc['Total Shareholders Funds']), df.loc['Total Shareholders Funds'],

                          df.loc['Total Revenue'], df.loc['Cost Of Materials Consumed'],
                          df.loc['Operating And Direct Expenses'],
                          df.loc['Employee Benefit Expenses'],
                          df.loc['Depreciation And Amortisation Expenses'],
                          df.loc['Other Expenses'],
                          df.loc['Total Expenses'] - df.loc['Finance Costs'],
                          df.loc['Profit/Loss Before Tax'] + df.loc['Finance Costs'],
                          df.loc['Finance Costs'], df.loc['Profit/Loss Before Tax'],
                          df.loc['Total Tax Expenses'], df.loc['Profit/Loss For The Period'],
                          df.loc['Minority Interest'],
                          df.loc['Net CashFlow From Operating Activities'],
                          df.loc['Cash And Cash Equivalents End Of Year'], capex.loc['Capex']])

    else:
        minority_interest = pd.DataFrame([0, 0, 0, 0, 0]).transpose()
        minority_interest.index = ['Minority Interest']
        minority_interest.columns = df_column

        df_val_data = \
            pd.DataFrame([df.loc['Inventories'], df.loc['Trade Receivables'],
                          df.loc['Trade Payables'],
                          (df.loc['Inventories'] + df.loc['Trade Receivables'] - \
                           df.loc['Trade Payables']) -
                          (df.loc['Inventories'].shift(-1) + df.loc['Trade Receivables'].shift(-1) \
                           - df.loc['Trade Payables'].shift(-1)),
                          df.loc['Long Term Borrowings'], df.loc['Short Term Borrowings'],
                          df.loc['Long Term Borrowings'] + df.loc['Short Term Borrowings'],
                          (df.loc['Total Capital And Liabilities'] -
                           df.loc['Total Shareholders Funds']), df.loc['Total Shareholders Funds'],

                          df.loc['Total Revenue'], df.loc['Cost Of Materials Consumed'],
                          df.loc['Operating And Direct Expenses'],
                          df.loc['Employee Benefit Expenses'],
                          df.loc['Depreciation And Amortisation Expenses'],
                          df.loc['Other Expenses'],
                          df.loc['Total Expenses'] - df.loc['Finance Costs'],
                          df.loc['Profit/Loss Before Tax'] + df.loc['Finance Costs'],
                          df.loc['Finance Costs'], df.loc['Profit/Loss Before Tax'],
                          df.loc['Total Tax Expenses'], df.loc['Profit/Loss For The Period'],
                          minority_interest.loc['Minority Interest'],

                          df.loc['Net CashFlow From Operating Activities'],
                          df.loc['Cash And Cash Equivalents End Of Year'], capex.loc['Capex']])

    df_val_data.index = ['Inventory', 'Accounts Receivables', 'Accounts Payable', 'Change in WC',
                         'Long Term Borrowings', 'Short Term Borrowings', 'Total Debt',
                         'Total Liability', 'Total Shareholders Equity',
                         'Total Revenue', 'COGS', 'Operating & Direct Expense',
                         'Employee Expenses', 'Depreciation And Amortisation', 'Other Expense',
                         'Total Expense', 'EBIT', 'Interest Expense', 'PBT', 'Tax Expense', 'PAT',
                         'Minority Interest', 'Cash Flow from Operations',
                         'Cash & cash equivalents', 'Capex']

    workbook = load_workbook(filename)
    sheet = workbook['Input Financials']

    cell_column = list(string.ascii_uppercase)
    for i in df_val_data.index:
        df1 = df_val_data.loc[i].tolist()
        table_row = cell_column[0] + str(3 + df_val_data.index.get_loc(i))
        sheet[table_row].value = i

        for j, val in enumerate(df1):
            if df_val_data.index.get_loc(i) == 0:
                table_column = cell_column[j + 1] + str(2)
                sheet[table_column] = df_val_data.columns.tolist()[j]
            cell = cell_column[j + 1] + str(3 + df_val_data.index.get_loc(i))
            sheet[cell].value = val
    workbook.save(filename)
    return df_val_data

def er_data_input(company, code, statement, capex, filename):
    ''' Takes company specific website url related details for data scraping
        Return financial statements '''

    ur_address = 'https://www.moneycontrol.com/financials/'
    if statement == 1:
        ur_pl = '/consolidated-profit-lossVI/'
        ur_bs = '/consolidated-balance-sheetVI/'
        ur_cf = '/consolidated-cash-flowVI/'

    else:
        ur_pl = '/profit-lossVI/'
        ur_bs = '/balance-sheetVI/'
        ur_cf = '/cash-flowVI/'

    url_pl = ur_address + company + ur_pl + code + '#' + code
    url_bs = ur_address + company + ur_bs + code + '#' + code
    url_cf = ur_address + company + ur_cf + code + '#' + code
    print('Financial Data taken from url: \n', url_bs, '\n', url_pl, '\n', url_cf)

    url_company = [url_pl, url_bs, url_cf]
    df_fs = get_financial_statements(url_company)
    df_pl = df_fs[0]
    df_bs = df_fs[1]
    df_cf = df_fs[2]
    df_equity_data = get_equity_research_data(df_bs, df_pl, df_cf, capex, statement, filename)
    return df_equity_data

def country_default_rating(country_rating):
    ''' Takes Moody's rating for country
        Returns country's default spread '''

    country_default = \
        {'moody rating': ['Aaa', 'Aa1', 'Aa2', 'Aa3', 'A1', 'A2', 'A3', 'Baa1', 'Baa2',
                          'Baa3', 'Ba1', 'Ba2', 'Ba3', 'B1', 'B2', 'B3', 'Caa1', 'Caa2',
                          'Caa3', 'Ca', 'C'],
         'spread': [0, 0.47, 0.58, 0.71, 0.83, 1.00, 1.41, 1.87, 2.23, 2.58, 2.93, 3.53,
                    4.22, 5.28, 6.46, 7.63, 8.80, 10.57, 11.73, 14.08, 17.50]
        }

    df_crp = pd.DataFrame(country_default)
    df1 = df_crp[df_crp['moody rating'] == country_rating]
    crp = df1['spread'].values.tolist()[0]
    print('Country Risk Premium :', crp, '%')
    return crp

def risk_free_rate(r_10y_tbond, crp):
    ''' Takes government 10 year T-Bond rates
        Returns country's risk free rate '''

    risk_free = r_10y_tbond - crp
    print('Risk Free Rate (rf) : ', round(risk_free, 2), '%')
    return risk_free

def company_default_rating(interest_coverage):
    ''' Takes Interest Coverage Ratio of comapny
        Returns company's default spread '''

    company_default = {'lower bound': [12.5, 9.5, 7.5, 6, 4.5, 4, 3.5, 3, 2.5, 2, 1.5,
                                       1.25, 0.8, 0.5, -100000],
                       'upper bound': [100000, 12.4999, 9.4999, 7.4999, 5.9999, 4.4999,
                                       3.9999, 3.4999, 2.9999, 2.4999, 1.9999, 1.4999,
                                       1.2499, 0.7999, 0.4999],
                       'spread': [0.63, 0.78, 0.98, 1.08, 1.22, 1.56, 2.00, 2.40, 3.51,
                                  4.21, 5.15, 8.20, 8.64, 11.34, 15.12]
                       }
    df_company_spread = pd.DataFrame(company_default)
    df1 = df_company_spread[df_company_spread['lower bound'] < interest_coverage]
    df2 = df1[df1['upper bound'] > interest_coverage]

    company_default_spread = df2['spread'].values.tolist()
    print('Company default spread : ', company_default_spread[0], '%')
    return company_default_spread[0]

def cost_of_debt(risk_free, crp, df):
    ''' Takes risk free rate, country default spread and financial data
        Returns Cost of Debt (kd) for Company '''

    ic_ratio = df.loc['EBIT'] / df.loc['Interest Expense']
    interest_coverage = ic_ratio[0]
    default_company = company_default_rating(interest_coverage)
    cost_debt = risk_free + crp + default_company
    print('Cost of Debt (kd) : ', round(cost_debt, 3), '%')
    print('-----------------------------------------------------')
    return cost_debt, default_company

def beta(stock, index, avg_de, avg_beta, tax_rate, df):
    ''' Takes Yahoo Finance ticker for stock and index, industry average beta and
        debt to equity ratio, tax rate and financial data
        Returns Bottom-Up beta and Historical beta of company '''

    # Importing historical data from Yahoo Finance using API - yahoo_fin.stock_info
    try:
        df1 = si.get_data(stock, '2015-01-01', None, True, "1wk")
    except AssertionError:
        stock = stock.replace('.NS', '.BO')
        df1 = si.get_data(stock, '2015-01-01', None, True, "1wk")
    print('Historical data taken from Yahoo Finance')
    print("The Stock Code: ", stock)

    df1.index.names = ['date']
    df1.index = [x.date() for x in df1.index]
    date_stock = df1.index.tolist()[0]

    df3 = si.get_data(index, '2015-01-01', None, True, "1wk")
    df3.index.names = ['date']
    df3.index = [x.date() for x in df3.index]
    date_index = df3.index.tolist()[0]

    if date_stock != date_index:
        df2 = si.get_data(index, date_stock, None, True, "1wk")
    else:
        df2 = df3

    df1['lg_PD'] = (np.log((df1['adjclose'].shift(-1) / df1['adjclose']))) * 100
    df2['lg_PD'] = (np.log((df2['adjclose'].shift(-1) / df2['adjclose']))) * 100

    # Calculation of Beta
    variance = df2["lg_PD"].var()
    covariance = df2["lg_PD"].cov(df1["lg_PD"])
    hist_beta = covariance / variance
    print("Historical Beta : ", round(hist_beta, 3))

    lever_beta = avg_beta / (1 + ((1 - tax_rate) * avg_de))
    if lever_beta == 0:
        bottom_beta = hist_beta
    else:
        de_ratio = df.loc['Total Debt'] / df.loc['Total Shareholders Equity']
        de_ratio_lastyear = de_ratio[0]
        bottom_beta = lever_beta * (1 + (1 - tax_rate) * de_ratio_lastyear)
        print('Bottom-Up Beta : ', round(bottom_beta, 3))
    return bottom_beta, hist_beta

def equity_risk_premium(erp_mature, df_rev):
    ''' Takes mature country's Equity risk premium and geography-wise revenue
        Returns Equity risk premium of country '''

    df_rev['percent_revenue'] = df_rev['Revenue'] / df_rev['Revenue'].sum()
    df_rev['country_erp'] = df_rev['percent_revenue'] * df_rev['CRP']
    erp = df_rev['country_erp'].sum() + erp_mature
    print('Equity Risk Premium :', round(erp, 3), '%')
    return erp

def cost_of_equity(erp_mature, risk_free, stock, index, avg_de, avg_beta, tax_rate, df, df_rev):
    ''' Takes mature country ERP, company stock ticker, industry average historical beta
        and debt to equity ratio and financial data
        Returns Cost of Equity (ke) for company '''

    bottom_beta = beta(stock, index, avg_de, avg_beta, tax_rate, df)
    erp = equity_risk_premium(erp_mature, df_rev)

    cost_equity = risk_free + bottom_beta[0] * erp
    print('Cost of Equity (ke) : ', round(cost_equity, 3), '%')
    return cost_equity, bottom_beta[0], erp, bottom_beta[1]

def cost_of_capital(cost_equity, cost_debt, tax_rate, df, debt_duration, df_a):
    ''' Takes kd, ke, tax rate, duration of debt instruments and financial data
        Returns Weighted Average Cost of Capital (WACC) for company '''

    df1 = df.loc['Total Debt']
    df2 = df.loc['Interest Expense']
    market_value_debt = np.pv(cost_debt / 100, round(debt_duration), -df2[0], -df1[0], 'begin')

    print('Market Value of Debt = ', market_value_debt)
    debt_equity_ratio = market_value_debt / df.loc['Total Shareholders Equity'][0]
    wacc = cost_debt * (1 - tax_rate) * debt_equity_ratio + cost_equity * (1 - debt_equity_ratio)
    print('WACC =', round(wacc, 3), '%')
    wacc = wacc + df_a['Change in WACC']
    print('-----------------------------------------------------')
    return wacc

def value_of_firm(df, tax_rate, risk_free, wacc, lease, df_a):
    ''' Takes WACC, risk free rate, tax rate, lease commitments, financial data
        and assumptions provided in input sheet of excel
        Returns Intrinsic Value of company '''

    # Calculating FCFF for previous years
    df1 = pd.DataFrame([df.loc['EBIT'], df.loc['EBIT'] * (1 - tax_rate),
                        df.loc['Depreciation And Amortisation'], df.loc['Capex'],
                        df.loc['Change in WC'],

                        df.loc['EBIT'] * (1 - tax_rate) \
                                + df.loc['Depreciation And Amortisation'] - df.loc['Capex'] \
                                + df.loc['Change in WC'],
                        df.loc['EBIT'] / df.loc['Total Revenue'],
                        df.loc['Depreciation And Amortisation'] / df.loc['Total Revenue'],
                        df.loc['Capex'] / df.loc['Total Revenue'],
                        df.loc['Change in WC'] / df.loc['Total Revenue'],
                        (df.loc['Capex'] + df.loc['Change in WC']) / (df.loc['EBIT'] * \
                                                                      (1 - tax_rate)),
                        df.loc['EBIT'] * (1 - tax_rate) / (df.loc['Total Debt'] + \
                                                           df.loc['Total Shareholders Equity'] - \
                                                           df.loc['Cash & cash equivalents']),
                        (df.loc['Capex'] + df.loc['Change in WC']) /
                        (df.loc['EBIT'] * (1 - tax_rate)) * \
                        df.loc['EBIT'] * (1 - tax_rate) / (df.loc['Total Debt'] +
                                                           df.loc['Total Shareholders Equity'] - \
                                                           df.loc['Cash & cash equivalents'])])

    df1.index = ['EBIT', 'EBIT*(1-t)', 'Depreciation And Amortisation', 'Capex',
                 'Change in WC', 'FCFF', 'Operating Margin', 'Dep & Amort (% Revenue)',
                 'Capex (% Revenue)', 'Change in WC (% Revenue)', 'Reinvestment rate',
                 'Return on Invested Capital (ROIC)', 'Growth (EBIT)']

    df3 = df1.loc['Growth (EBIT)'].tolist()
    growth_ebit = (df3[0] + df3[1] + df3[2]) / 3
    print('Average historical EBIT growth: ', round(growth_ebit * 100, 3), '%')

    df4 = df1.loc['Dep & Amort (% Revenue)'].tolist()
    dep_and_amort = (df4[0] + df4[1] + df4[2]) / 3

    df5 = df1.loc['Change in WC (% Revenue)'].tolist()
    work_capital = (df5[0] + df5[1] + df5[2]) / 3

    df6 = df1.loc['Capex (% Revenue)'].tolist()
    cap_ex = (df6[0] + df6[1] + df6[2]) / 3

    #df7 = df1.loc['Operating Margin'].tolist()
    #oper_margin = (df7[0] + df7[1] + df7[2]) / 3

    # Growth Story for the company
    growth = pd.DataFrame([df_a['EBIT (x times)'], growth_ebit * df_a['EBIT (x times)'],
                           cap_ex * df_a['Change in Capex (% of  Revenue)'],
                           dep_and_amort * df_a['Change in Dep & Amort (% of Revenue)'],
                           work_capital * df_a['Change in WC (% of Revenue)']])

    growth.index = ['EBIT (x times)', 'Growth in EBIT(%)', 'Capex (% of Revenue)',
                    'Dep & Amort (% Revenue)', 'Change in WC (% Revenue)']

    growth.loc['Growth in EBIT(%)', 0] = df3[0]
    growth.loc['Dep & Amort (% Revenue)', 0] = df4[0]
    growth.loc['Change in WC (% Revenue)', 0] = df5[0]
    growth.loc['Capex (% of Revenue)', 0] = df6[0]

    ebit = []
#    margin = []
    rev = []
    dep = []
    capex = []
    working_capital = []
    fcff = []
    present_value = []

    for i in growth.columns:
        if i == 0:
            ebit.append(df.loc['EBIT'].tolist()[0])
            #           margin.append(oper_margin)
            rev.append(df.loc['Total Revenue'].tolist()[0])
            dep.append(df.loc['Depreciation And Amortisation'].tolist()[0])
            working_capital.append(df.loc['Change in WC'].tolist()[0])
            capex.append(df.loc['Capex'].tolist()[0])

        else:
            ebit.append(ebit[-1] * (1 + growth.loc['Growth in EBIT(%)', i]))
            #           margin.append(oper_margin)
            rev.append(rev[-1] * (1 + growth.loc['Growth in EBIT(%)', i]))
            dep.append(rev[-1] * growth.loc['Dep & Amort (% Revenue)', i])
            capex.append(rev[-1] * growth.loc['Capex (% of Revenue)', i])
            working_capital.append(rev[-1] * growth.loc['Change in WC (% Revenue)', i])

        fcff.append(ebit[-1] * (1 - tax_rate) + dep[-1] - capex[-1] - working_capital[-1])

    # Terminal Value Calculations
    fcff[-1] = fcff[-1] * (1 + risk_free / 100) / (wacc[len(wacc)-1] / 100 - risk_free / 100)

    for i, val in enumerate(fcff):
        temp = fcff[i]
        for j in range(i, 0, -1):
            temp = np.pv(wacc[j] / 100, 1, -temp, 0)
        present_value.append(temp)

    df8 = pd.DataFrame([ebit, growth.loc['Growth in EBIT(%)'], rev,
                        growth.loc['Dep & Amort (% Revenue)'], dep,
                        growth.loc['Capex (% of Revenue)'], capex,
                        growth.loc['Change in WC (% Revenue)'], working_capital,
                        fcff, wacc/100, present_value])

    df8.index = ['EBIT', 'Growth in EBIT(%)', 'Total Revenue', 'Dep & Amort (% Revenue)',
                 'Depreciation And Amortisation', 'Capex (% Revenue)', 'Capex',
                 'Change in WC (% Revenue)', 'Change in WC', 'FCFF', 'WACC', 'PV']

    # Renaming column names
    last_year = df1.columns[0][-2:]
    month = df1.columns[0][:-2]
    for i, val in enumerate(df8.columns):
        if i == len(df8.columns) - 1:
            df8 = df8.rename(columns={val: 'Terminal Value'})
        if i == 0:
            temp = month + str(int(last_year) + i) + ' (A)'
        else:
            temp = month + str(int(last_year) + i) + ' (E)'
        df8 = df8.rename(columns={val: temp})

    value = df8.loc['PV'].sum() - df.loc['Total Debt'].tolist()[0] + \
            df.loc['Cash & cash equivalents'].tolist()[0] - \
            abs(df.loc['Minority Interest'].tolist()[0]) - lease

    print('Value of firm (in Crs) :', round(value, 3))
    return value, growth_ebit, df8

def current_stock_price(stock, value):
    ''' Takes company's Yahoo Finance ticker and intrinsic value of company
        Returns Value per share '''

    df = si.get_quote_table(stock)
    cmp = df['Quote Price']
    market_cap = df['Market Cap']
    if market_cap[-1] == 'T':
        market_cap = float(market_cap[:-1]) * 10 ** 12 / 10 ** 7
    elif market_cap[-1] == 'B':
        market_cap = float(market_cap[:-1]) * 10 ** 9 / 10 ** 7
    elif market_cap[-1] == 'M':
        market_cap = float(market_cap[:-1]) * 10 ** 6 / 10 ** 7
    else:
        market_cap = float(market_cap)
    print('Market Capitalisation (in Crs):', market_cap)
    share_out = market_cap / cmp
    print('Shares Outstanding (in Crs):', round(share_out, 3))
    print('-----------------------------------------------------')
    print('CMP : Rs ', round(cmp, 3))
    value_per_share = value / share_out
    print('Value per share: Rs ', round(value_per_share, 3))
    print('-----------------------------------------------------')

    if value_per_share <= cmp:
        print('The Stock is OVER PRICED')
    else:
        print('The Stock is UNDER PRICED')
    return value_per_share, market_cap, share_out, cmp

def intrinsic_valuation(filename):
    ''' Takes excel file path
        Calls other functions
        Return key financial data, forecasted FCFF and forcasted P&L statement'''

    df1 = importing_input_data(filename)
    df_f = df1[0]
    df_a = df1[1]
    df_rev = df1[2]

    r_10y_tbond = df_f['T-Bond Rates (10 Years)']
    country_rating = df_f["Country Rating (Moody's)"]
    erp_mature = df_f['ERP (Mature)']
    stock = df_f['Stock']
    index = df_f['Index']
    tax_rate = df_f['Tax Rate (%)'] / 100
    debt_duration = df_f['Average duration of debt']
    lease = df_f['Lease Liability']
    avg_de = df_f['Average Debt/Equity Ratio']
    avg_beta = df_f['Average Beta']

    capex = [df_f['Capex (T)'], df_f['Capex (T-1)'], df_f['Capex (T-2)'], 0, 0]
    capex = pd.DataFrame(capex).transpose()
    capex.index = ['Capex']

    print('INTRINSIC VALUATION')
    print('-----------------------------------------------------')
    print('Company :', df_f['Company'])
    print('T-Bond Rates (10 Years) :', r_10y_tbond, '%')
    print('ERP (Mature Country) :', erp_mature, '%')
    print('Tax Rate (%) :', tax_rate * 100, '%')
    print('-----------------------------------------------------')

    # Input for Financial Statements Scrapping from Money Control
    company = df_f['Code 1']
    code = df_f['Code 2']
    # Enter '0' for standalone and '1' for consolidated
    statement = df_f['Statement (0: S & 1:C)']

    # For getting the financial statements
    df = er_data_input(company, code, statement, capex, filename)
    print('-----------------------------------------------------')

    # Calculating Risk Free Premium and Country Default Risk
    crp = country_default_rating(country_rating)
    risk_free = risk_free_rate(r_10y_tbond, crp)

    # Calculating Weighted average Cost of Capital
    cost_debt = cost_of_debt(risk_free, crp, df)
    cost_equity = cost_of_equity(erp_mature, risk_free, stock, index, \
                                 avg_de, avg_beta, tax_rate, df, df_rev)
    wacc = cost_of_capital(cost_equity[0], cost_debt[0], tax_rate, df, debt_duration, df_a)
    value = value_of_firm(df, tax_rate, risk_free, wacc, lease, df_a)
    price = current_stock_price(stock, value[0])

    # Exporting Valuation Data to Output sheet of Intrinsic_Valuation.xlsx
    workbook = load_workbook(filename)
    sheet = workbook['Output Sheet']

    sheet['B2'].value = df_f['Company']
    sheet['B3'].value = risk_free
    sheet['B4'].value = crp
    sheet['B5'].value = cost_debt[1]
    sheet['B6'].value = cost_equity[3]
    sheet['B7'].value = cost_equity[1]
    sheet['B8'].value = cost_equity[2]
    sheet['B9'].value = cost_debt[0]
    sheet['B10'].value = cost_equity[0]
    sheet['B11'].value = wacc[0]
    sheet['B12'].value = value[1] * 100
    sheet['B13'].value = value[0]
    sheet['B14'].value = price[1]
    sheet['B15'].value = price[2]
    sheet['B16'].value = price[3]
    sheet['B17'].value = price[0]

    if price[3] > price[0]:
        sheet['B18'].value = 'Overpriced'
    else:
        sheet['B18'].value = 'Underpriced'

    cell_column = list(string.ascii_uppercase)
    df1 = value[2]
    for i in df1.index:
        df2 = df1.loc[i].tolist()
        table_row = cell_column[0] + str(22 + df1.index.get_loc(i))
        sheet[table_row].value = i

        for j, val in enumerate(df2):
            if df1.index.get_loc(i) == 0:
                table_column = cell_column[j + 1] + str(21)
                sheet[table_column] = df1.columns.tolist()[j]

            cell = cell_column[j + 1] + str(22 + df1.index.get_loc(i))
            sheet[cell].value = val
    workbook.save(filename)
    return price

# INPUT SHEET FOR THE CODE
FILENAME = r'C:\Users\Student\Desktop\2k20\Valuation\0. Intrinsic Valuation_Input Sheet\Valuation_InputOutputSheet_R0.xlsx'
intrinsic_valuation(FILENAME)